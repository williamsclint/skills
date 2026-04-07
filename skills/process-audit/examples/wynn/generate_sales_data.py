"""
generate_sales_data.py
Generates 30 days of fictitious daily sales data for the top 20 Wynn retail stores.
Outputs a formatted Excel file.
"""

import random
import datetime
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

random.seed(42)

# ── Store definitions ─────────────────────────────────────────────────────────
# (name, property, product_category, weekday_min, weekday_max)
# Revenue ranges are realistic estimates for luxury Strip retail
STORES = [
    ("Graff",                       "Wynn Esplanade",   "Fine Jewellery",                          35_000,  220_000),
    ("Bijan",                       "Wynn Esplanade",   "Menswear",                                12_000,   90_000),
    ("Van Cleef & Arpels",          "Wynn Esplanade",   "Fine Jewellery & Watches",                55_000,  210_000),
    ("Cartier",                     "Wynn Esplanade",   "Fine Jewellery & Watches",                75_000,  260_000),
    ("Hermes",                      "Encore Esplanade", "Clothing, Leather Goods & Accessories",   85_000,  290_000),
    ("Chanel (Fine Jewellery)",     "Wynn Esplanade",   "Fine Jewellery & Watches",                65_000,  195_000),
    ("Chanel (Fashion)",            "Wynn Esplanade",   "Clothing & Accessories",                  95_000,  330_000),
    ("BVLGARI",                     "Wynn Esplanade",   "Fine Jewellery & Watches",                45_000,  160_000),
    ("Dior",                        "Wynn Esplanade",   "Clothing & Accessories",                  85_000,  270_000),
    ("Brioni",                      "Wynn Esplanade",   "Menswear",                                18_000,   75_000),
    ("Louis Vuitton",               "Wynn Esplanade",   "Leather Goods, Clothing & Accessories",  175_000,  460_000),
    ("Prada",                       "Wynn Esplanade",   "Clothing & Accessories",                  75_000,  230_000),
    ("Loewe",                       "Wynn Plaza",       "Clothing & Leather Goods",                40_000,  135_000),
    ("Celine",                      "Wynn Plaza",       "Clothing & Leather Goods",                48_000,  145_000),
    ("Saint Laurent",               "Wynn Plaza",       "Clothing & Accessories",                  60_000,  175_000),
    ("Gucci",                       "Wynn Esplanade",   "Clothing & Accessories",                  95_000,  290_000),
    ("Givenchy",                    "Wynn Esplanade",   "Clothing & Accessories",                  30_000,   88_000),
    ("Buccellati",                  "Wynn Esplanade",   "Fine Jewellery",                          22_000,  105_000),
    ("Oscar de la Renta",           "Wynn Esplanade",   "Clothing & Accessories",                  22_000,   78_000),
    ("Rolex",                       "Wynn Esplanade",   "Watches",                                 85_000,  310_000),
]

# ── Date range: today back 30 days ────────────────────────────────────────────
TODAY = datetime.date(2026, 3, 26)
DATES = [TODAY - datetime.timedelta(days=i) for i in range(30)]  # 26 Mar → 25 Feb

# ── Build rows ────────────────────────────────────────────────────────────────
rows = []
for store in STORES:
    name, prop, category, low, high = store
    for day in DATES:
        # Weekends ~40% higher, Fridays ~20% higher
        multiplier = 1.0
        if day.weekday() == 4:   # Friday
            multiplier = 1.20
        elif day.weekday() in (5, 6):  # Sat/Sun
            multiplier = 1.40

        # Add some random variance
        revenue = int(random.uniform(low, high) * multiplier)
        # Round to nearest $500
        revenue = round(revenue / 500) * 500

        rows.append({
            "Vendor Name":      name,
            "Trading Day":      day,
            "Revenue (USD)":    revenue,
            "Product Category": category,
            "Wynn Property":    prop,
            "Report Submitted": "Yes",   # default; overwritten below
        })

# ── Set 5% of rows to Report Submitted = No, randomly ────────────────────────
total_rows = len(rows)           # 600
no_count   = round(total_rows * 0.05)   # 30
no_indices = random.sample(range(total_rows), no_count)
for i in no_indices:
    rows[i]["Report Submitted"] = "No"

# ── Sort: Vendor → Date ───────────────────────────────────────────────────────
rows.sort(key=lambda r: (r["Vendor Name"], r["Trading Day"]))

# ── Excel output ──────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Daily Sales Data"

# Colours
NAVY        = "1A1A2E"
GOLD        = "B89655"
LIGHT_GOLD  = "F7F0E0"
WHITE       = "FFFFFF"
RED_FILL    = "FDECEA"
RED_FONT    = "C0392B"
GREEN_FILL  = "EAF7EC"
GREEN_FONT  = "1E8449"
STRIPE_A    = "FAFAFA"
STRIPE_B    = "FFFFFF"

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def thin_border():
    s = Side(style="thin", color="DDDDDD")
    return Border(left=s, right=s, top=s, bottom=s)

# ── Title row ─────────────────────────────────────────────────────────────────
ws.merge_cells("A1:F1")
title_cell = ws["A1"]
title_cell.value = "WYNN RESORTS — RETAIL DAILY SALES DATA  |  FEB 25 – MAR 26, 2026"
title_cell.font      = Font(name="Garamond", bold=True, size=13, color=WHITE)
title_cell.fill      = fill(NAVY)
title_cell.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 28

# ── Sub-title row ─────────────────────────────────────────────────────────────
ws.merge_cells("A2:F2")
sub_cell = ws["A2"]
sub_cell.value = "Top 20 Retail Boutiques  |  Fictitious sales data for demonstration purposes only"
sub_cell.font      = Font(name="Garamond", italic=True, size=9, color="6B6B6B")
sub_cell.fill      = fill(LIGHT_GOLD)
sub_cell.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[2].height = 16

# ── Header row ────────────────────────────────────────────────────────────────
headers = ["Vendor Name", "Trading Day", "Revenue (USD)", "Product Category",
           "Wynn Property", "Report Submitted"]
for col_idx, header in enumerate(headers, start=1):
    cell = ws.cell(row=3, column=col_idx, value=header)
    cell.font      = Font(name="Garamond", bold=True, size=10, color=WHITE)
    cell.fill      = fill(NAVY)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = thin_border()
ws.row_dimensions[3].height = 20

# ── Data rows ─────────────────────────────────────────────────────────────────
for row_idx, row in enumerate(rows, start=4):
    stripe = STRIPE_A if row_idx % 2 == 0 else STRIPE_B

    values = [
        row["Vendor Name"],
        row["Trading Day"],
        row["Revenue (USD)"],
        row["Product Category"],
        row["Wynn Property"],
        row["Report Submitted"],
    ]

    for col_idx, value in enumerate(values, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border()

        # Base styling
        cell.font      = Font(name="Garamond", size=9)
        cell.alignment = Alignment(vertical="center")

        # Report Submitted colouring
        submitted = row["Report Submitted"]
        if col_idx == 6:
            if submitted == "No":
                cell.fill = fill(RED_FILL)
                cell.font = Font(name="Garamond", size=9, bold=True, color=RED_FONT)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.fill = fill(GREEN_FILL)
                cell.font = Font(name="Garamond", size=9, color=GREEN_FONT)
                cell.alignment = Alignment(horizontal="center", vertical="center")
        else:
            cell.fill = fill(stripe)

        # Revenue: currency format
        if col_idx == 3:
            cell.number_format = '"$"#,##0'
            cell.alignment = Alignment(horizontal="right", vertical="center")

        # Date format
        if col_idx == 2:
            cell.number_format = "DD-MMM-YYYY"
            cell.alignment = Alignment(horizontal="center", vertical="center")

# ── Column widths ─────────────────────────────────────────────────────────────
col_widths = [28, 16, 16, 38, 20, 18]
for i, width in enumerate(col_widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = width

# ── Freeze header ─────────────────────────────────────────────────────────────
ws.freeze_panes = "A4"

# ── Auto-filter ───────────────────────────────────────────────────────────────
ws.auto_filter.ref = f"A3:F{3 + len(rows)}"

# ── Summary sheet ─────────────────────────────────────────────────────────────
ws2 = wb.create_sheet("Summary by Vendor")

ws2.merge_cells("A1:D1")
s_title = ws2["A1"]
s_title.value = "WYNN RESORTS — 30-DAY REVENUE SUMMARY BY VENDOR"
s_title.font      = Font(name="Garamond", bold=True, size=12, color=WHITE)
s_title.fill      = fill(NAVY)
s_title.alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 24

sum_headers = ["Vendor Name", "Total Revenue (30 Days)", "Days Not Reported", "Avg Daily Revenue"]
for ci, h in enumerate(sum_headers, 1):
    c = ws2.cell(row=2, column=ci, value=h)
    c.font      = Font(name="Garamond", bold=True, size=10, color=WHITE)
    c.fill      = fill(NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border    = thin_border()
ws2.row_dimensions[2].height = 20

# Aggregate
from collections import defaultdict
totals      = defaultdict(int)
missing     = defaultdict(int)

for row in rows:
    totals[row["Vendor Name"]] += row["Revenue (USD)"]
    if row["Report Submitted"] == "No":
        missing[row["Vendor Name"]] += 1

# Sort by total revenue descending
sorted_vendors = sorted(totals.items(), key=lambda x: x[1], reverse=True)

for ri, (vendor, total) in enumerate(sorted_vendors, start=3):
    avg = total // 30
    miss = missing.get(vendor, 0)
    stripe = STRIPE_A if ri % 2 == 0 else STRIPE_B

    vals = [vendor, total, miss, avg]
    for ci, val in enumerate(vals, 1):
        c = ws2.cell(row=ri, column=ci, value=val)
        c.font   = Font(name="Garamond", size=9)
        c.border = thin_border()
        c.fill   = fill(stripe)
        c.alignment = Alignment(vertical="center")
        if ci in (2, 4):
            c.number_format = '"$"#,##0'
            c.alignment = Alignment(horizontal="right", vertical="center")
        if ci == 3 and val > 0:
            c.fill = fill(RED_FILL)
            c.font = Font(name="Garamond", size=9, bold=True, color=RED_FONT)
            c.alignment = Alignment(horizontal="center", vertical="center")

ws2.column_dimensions["A"].width = 28
ws2.column_dimensions["B"].width = 22
ws2.column_dimensions["C"].width = 18
ws2.column_dimensions["D"].width = 22

# ── Save ──────────────────────────────────────────────────────────────────────
import os
out_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "wynn-retail-sales-30day.xlsx")
wb.save(out_path)
print(f"Saved -> {out_path}")
print(f"Total rows: {len(rows)} | Rows with Report Submitted = No: {no_count}")
